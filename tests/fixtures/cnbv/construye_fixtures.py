"""Reduce los boletines reales de la CNBV a fixtures versionables.

Los originales pesan 2.4 MB (banca) y 475 KB (SOFIPOs) y traen treinta y tres
hojas de las que se leen tres. Este script deja **las mismas hojas, las mismas
filas de encabezado y los mismos valores**, recortando sólo el número de
instituciones y las hojas que no se usan.

Los fixtures siguen siendo datos reales —el criterio de aceptación de §8 exige
que no sean sintéticos—, sólo que menos. Que el recorte sea reproducible con
este script es lo que hace verificable esa afirmación.

Una salvedad: al reescribir el libro, openpyxl reserializa los flotantes y el
último dígito puede cambiar (`1.6777185356184765` sale `1.677718535618477`).
Los tests comparan a cuatro decimales, que es la precisión con la que la base
guarda un porcentaje, así que da igual — pero conviene saberlo antes de
perseguir una diferencia que no está en el parser.

Uso, con los originales descargados en el directorio que se le pase:

    python tests/fixtures/cnbv/construye_fixtures.py <directorio-con-originales>
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

AQUI = Path(__file__).parent

#: `(original, salida, {hoja: filas a conservar})`. El tope de filas se cuenta
#: desde la 1, así que incluye encabezados: hay que dejar bastantes para que
#: aparezcan las instituciones del catálogo.
RECORTES: tuple[tuple[str, str, dict[str, int]], ...] = (
    (
        "BE BM 202605.xlsx",
        "banca_202605.xlsx",
        {"CCT": 60, "Art_121": 60, "CaptRec": 60},
    ),
    (
        "BE_SOFIPOS_202603.xlsx",
        "sofipos_202603.xlsx",
        {"Sociedades_1": 52, "Sociedades_2": 52, "Tasas_Implícitas": 52},
    ),
)

#: Columnas a conservar. Sobran de largo para todos los conceptos declarados en
#: `fuentes.py` y mantienen los índices originales, que es lo que importa.
COLUMNAS = 20


def recortar(origen: Path, destino: Path, hojas: dict[str, int]) -> None:
    libro = load_workbook(origen, read_only=True, data_only=True)
    nuevo = Workbook()
    nuevo.remove(nuevo.active)  # type: ignore[arg-type]
    try:
        for nombre, tope in hojas.items():
            if nombre not in libro.sheetnames:
                raise SystemExit(f"{origen.name}: falta la hoja '{nombre}'")
            fuente = libro[nombre]
            hoja = nuevo.create_sheet(nombre)
            for fila in fuente.iter_rows(max_row=tope, max_col=COLUMNAS, values_only=True):
                hoja.append(list(fila))
    finally:
        libro.close()
    nuevo.save(destino)
    print(f"   {destino.name}: {destino.stat().st_size // 1024} KB")


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    origenes = Path(sys.argv[1])
    for original, salida, hojas in RECORTES:
        ruta = origenes / original
        if not ruta.exists():
            # También se acepta el nombre con guiones bajos, que es como queda
            # tras descargarlo en un sistema de archivos poco amigo de espacios.
            ruta = origenes / original.replace(" ", "_")
        if not ruta.exists():
            raise SystemExit(f"no encuentro {original} en {origenes}")
        recortar(ruta, AQUI / salida, hojas)


if __name__ == "__main__":
    main()
