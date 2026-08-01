"""Lectura de los boletines XLSX, con validación antes que confianza.

Un solo lector para banca múltiple y para SOFIPOs: las dos publicaciones tienen
formatos distintos, pero la diferencia está entera en la declaración de
`fuentes.py` —qué hoja, qué fila, qué encabezado— y no en el código que lee.
Dos parsers separados habrían divergido en la primera corrección.

**Se valida y luego se lee, nunca al revés.** Antes de tomar una cifra:

1. La hoja tiene que existir con ese nombre.
2. El encabezado de cada concepto tiene que decir lo que se espera.
3. Dentro del bloque de tres periodos, tiene que haber una columna cuyo periodo
   **coincida con el del boletín**.

Si algo de eso falla, se lanza `FormatoInesperado` con el detalle. Es el
criterio de §8: un cambio de formato rompe el job con un error explícito y
nunca carga datos malinterpretados. La alternativa —leer la columna de al lado
sin enterarse— publicaría un IMOR del año pasado como si fuera el de este mes,
y eso mueve banderas que la gente lee para decidir dónde pone su dinero.

Se usa `openpyxl` en modo `read_only` y **no pandas**: el scheduler del VPS
tiene 256 MB y estos libros traen treinta y tres hojas. El modo de sólo lectura
recorre por filas sin materializar el libro entero.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from core.logging import get_logger
from ingest_cnbv.fuentes import Concepto, Hoja

log = get_logger(__name__)

#: Cómo marca la CNBV lo que no aplica o no reportó. `CAME` aparece así en el
#: boletín de marzo de 2026. No es cero: es ausencia, y un cero mentiría.
_SIN_DATO = frozenset({"n.a.", "n.d.", "na", "nd", "-", "--", "", "*"})

MESES_ES: dict[str, int] = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


class FormatoInesperado(Exception):
    """El archivo no tiene la forma declarada. **Nunca se ingiere a medias.**"""


@dataclass(slots=True)
class FilaInstitucion:
    """Lo leído para una institución, con el nombre tal cual lo escribe la CNBV."""

    nombre_cnbv: str
    valores: dict[str, Decimal | str | None] = field(default_factory=dict)

    def numero(self, campo: str) -> Decimal | None:
        valor = self.valores.get(campo)
        return valor if isinstance(valor, Decimal) else None

    def texto(self, campo: str) -> str | None:
        valor = self.valores.get(campo)
        return valor if isinstance(valor, str) else None


def _normalizar(texto: object) -> str:
    """Minúsculas sin acentos ni espacios de más, para comparar encabezados.

    La CNBV escribe «Índice de morosidad 2/\\n(%)» y en otro periodo puede
    escribir «Indice de morosidad (%)». Comparar literalmente rompería el job
    por una tilde, que es exactamente el fallo ruidoso *equivocado*.
    """
    if texto is None:
        return ""
    crudo = str(texto).replace("\n", " ").strip().lower()
    sin_acentos = "".join(
        c for c in unicodedata.normalize("NFD", crudo) if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"\s+", " ", sin_acentos)


def _a_decimal(valor: object) -> Decimal | None:
    """La celda como número, o `None` si la CNBV marcó que no hay dato."""
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, int | float | Decimal):
        return Decimal(str(valor))
    texto = str(valor).strip().replace(",", "").replace("%", "")
    if _normalizar(texto) in _SIN_DATO:
        return None
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _periodo_de_celda(valor: object) -> tuple[int, int] | None:
    """`(año, mes)` de una celda de la fila de periodos.

    Banca múltiple las escribe como fechas de verdad (`2026-05-01`) y SOFIPOs
    como texto en español (`Marzo 2026`). Las dos conviven en el mismo formato
    de archivo, así que las dos se entienden aquí.
    """
    if isinstance(valor, datetime):
        return (valor.year, valor.month)
    if isinstance(valor, date):
        return (valor.year, valor.month)
    texto = _normalizar(valor)
    if not texto:
        return None
    coincidencia = re.match(r"([a-z]+)\s+(\d{4})", texto)
    if coincidencia:
        mes = MESES_ES.get(coincidencia.group(1))
        if mes:
            return (int(coincidencia.group(2)), mes)
    coincidencia = re.match(r"(\d{4})-(\d{2})", texto)
    if coincidencia:
        return (int(coincidencia.group(1)), int(coincidencia.group(2)))
    return None


def _columna_del_periodo(
    fila_periodos: tuple[Any, ...],
    concepto: Concepto,
    periodo: tuple[int, int],
    hoja: str,
) -> int:
    """Dentro del bloque del concepto, la columna del periodo del boletín.

    Buscar por periodo y no tomar «la tercera» es lo que impide el fallo
    silencioso: el bloque de IMOR de `CCT` empieza en la columna F, pero esa F
    es el mismo mes del **año anterior**. Cargarla sería publicar el IMOR de
    hace un año como si fuera el de este mes.
    """
    encontradas: list[tuple[int, tuple[int, int]]] = []
    for desplazamiento in range(concepto.ancho):
        indice = concepto.columna + desplazamiento
        celda = fila_periodos[indice - 1] if indice - 1 < len(fila_periodos) else None
        leido = _periodo_de_celda(celda)
        if leido is not None:
            encontradas.append((indice, leido))
        if leido == periodo:
            return indice

    raise FormatoInesperado(
        f"{hoja}: el concepto '{concepto.encabezado}' no tiene ninguna columna "
        f"del periodo {periodo[0]}-{periodo[1]:02d}. "
        f"Encontrado: {[(i, f'{a}-{m:02d}') for i, (a, m) in encontradas] or 'nada'}"
    )


def leer_hoja(
    ruta: Path, hoja: Hoja, *, periodo: date, fuente_url: str | None = None
) -> list[FilaInstitucion]:
    """Lee una hoja del boletín. Lanza `FormatoInesperado` si no cuadra."""
    from openpyxl import load_workbook

    try:
        libro = load_workbook(ruta, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl lanza de todo
        raise FormatoInesperado(
            f"{ruta.name}: no se pudo abrir como libro OOXML ({type(exc).__name__}). "
            f"Los boletines anteriores a 2016 vienen en .xls, que no se lee aquí."
        ) from exc

    try:
        if hoja.nombre not in libro.sheetnames:
            raise FormatoInesperado(
                f"{ruta.name}: no tiene la hoja '{hoja.nombre}'. "
                f"Hojas presentes: {', '.join(libro.sheetnames[:12])}"
            )
        ws = libro[hoja.nombre]
        filas = list(ws.iter_rows(values_only=True))
    finally:
        libro.close()

    esperado = (periodo.year, periodo.month)
    encabezados = _fila(filas, hoja.fila_encabezado, hoja.nombre, ruta)
    _validar_encabezados(encabezados, hoja, ruta)

    columnas: dict[str, int] = {}
    if hoja.fila_periodo is None:
        columnas = {c.campo: c.columna for c in hoja.conceptos}
    else:
        periodos = _fila(filas, hoja.fila_periodo, hoja.nombre, ruta)
        for concepto in hoja.conceptos:
            columnas[concepto.campo] = _columna_del_periodo(
                periodos, concepto, esperado, f"{ruta.name}::{hoja.nombre}"
            )

    resultado: list[FilaInstitucion] = []
    for fila in filas[hoja.fila_datos - 1 :]:
        nombre = _celda(fila, hoja.col_institucion)
        if nombre is None or not str(nombre).strip():
            continue
        etiqueta = str(nombre).strip()
        if _no_es_institucion(etiqueta, hoja):
            continue

        valores: dict[str, Decimal | str | None] = {}
        for campo, indice in columnas.items():
            crudo = _celda(fila, indice)
            numero = _a_decimal(crudo)
            # `categoria` es un nivel romano (I–V), no una cifra: se conserva
            # como texto. El resto se guarda como Decimal o como nada.
            valores[campo] = (
                str(crudo).strip() if numero is None and campo == "categoria" and crudo else numero
            )
        resultado.append(FilaInstitucion(nombre_cnbv=etiqueta, valores=valores))

    if not resultado:
        raise FormatoInesperado(
            f"{ruta.name}::{hoja.nombre}: no se leyó ninguna institución desde la "
            f"fila {hoja.fila_datos}. El formato cambió."
        )

    log.info(
        "cnbv_hoja_leida",
        archivo=ruta.name,
        hoja=hoja.nombre,
        instituciones=len(resultado),
        periodo=periodo.isoformat(),
        fuente_url=fuente_url,
    )
    return resultado


#: Una nota al pie empieza numerando: «4/ ICAP (Índice de Capitalización) =
#: Capital Neto / Activos...». Van en la misma columna que los nombres, debajo
#: de la tabla, y sin esto entran como cinco instituciones más.
_NOTA_AL_PIE = re.compile(r"^\d+\s*/")

#: Ninguna institución mexicana se llama con sesenta caracteres. Lo que los
#: pasa es prosa: encabezados de continuación, aclaraciones, fuentes.
_LARGO_MAXIMO_DE_NOMBRE = 60


def _no_es_institucion(etiqueta: str, hoja: Hoja) -> bool:
    """¿Esta fila es un agregado, una nota al pie o prosa suelta?"""
    normalizado = _normalizar(etiqueta)
    return (
        normalizado in hoja.agregados
        or normalizado.startswith("sistema")
        or bool(_NOTA_AL_PIE.match(etiqueta))
        or len(etiqueta) > _LARGO_MAXIMO_DE_NOMBRE
    )


def _fila(filas: list[tuple[Any, ...]], numero: int, hoja: str, ruta: Path) -> tuple[Any, ...]:
    if numero - 1 >= len(filas):
        raise FormatoInesperado(
            f"{ruta.name}::{hoja}: se esperaba contenido en la fila {numero} y la "
            f"hoja sólo tiene {len(filas)}."
        )
    return filas[numero - 1]


def _celda(fila: tuple[Any, ...], columna: int) -> Any:
    return fila[columna - 1] if columna - 1 < len(fila) else None


def _validar_encabezados(encabezados: tuple[Any, ...], hoja: Hoja, ruta: Path) -> None:
    for concepto in hoja.conceptos:
        leido = _normalizar(_celda(encabezados, concepto.columna))
        if not leido.startswith(_normalizar(concepto.encabezado)):
            raise FormatoInesperado(
                f"{ruta.name}::{hoja.nombre}: la columna {concepto.columna} debía "
                f"encabezar '{concepto.encabezado}' y dice '{leido or '(vacío)'}'. "
                f"El formato del boletín cambió; los parsers no adivinan."
            )


def combinar(*lecturas: list[FilaInstitucion]) -> dict[str, FilaInstitucion]:
    """Une varias hojas en una fila por institución, indexada por nombre CNBV.

    Los conceptos viven repartidos entre hojas —el IMOR en una, la captación en
    otra— y la tabla de destino los quiere juntos.
    """
    unidas: dict[str, FilaInstitucion] = {}
    for lectura in lecturas:
        for fila in lectura:
            actual = unidas.setdefault(
                fila.nombre_cnbv, FilaInstitucion(nombre_cnbv=fila.nombre_cnbv)
            )
            actual.valores.update(fila.valores)
    return unidas


__all__ = [
    "MESES_ES",
    "FilaInstitucion",
    "FormatoInesperado",
    "combinar",
    "leer_hoja",
]
